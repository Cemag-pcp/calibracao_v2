from django.shortcuts import render
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required
import tempfile
import os
from openpyxl.drawing.image import Image
import copy
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from cadastro.models import Funcionario, PontoCalibracao
from ficha.models import AssinaturaInstrumento,StatusInstrumento

# Create your views here.

@login_required
def ficha_por_funcionario(request):

    if request.method == 'GET':

        funcionarios = Funcionario.objects.all()

        return render(request, "ficha.html",{"funcionarios":funcionarios})

@login_required
def emissao_ficha_por_funcionario(request, id):
    if request.method == 'GET':
        funcionario = Funcionario.objects.filter(pk=id).first()

        if funcionario:
            # Converte o QuerySet em uma lista de dicionários
            assinatura_funcionario = AssinaturaInstrumento.objects.filter(assinante=funcionario).order_by('data_entrega')
            status_funcionario = StatusInstrumento.objects.filter(funcionario=funcionario).order_by('data_entrega')

            dados_combinados = []
            for assinatura in assinatura_funcionario:
                dados_combinados.append({
                    'data_entrega': assinatura.data_entrega,
                    'instrumento': assinatura.instrumento,
                    'motivo': assinatura.motivo,
                    'foto_assinatura': assinatura.foto_assinatura,
                    'tipo': 'Assinatura'
                })

            for status in status_funcionario:
                dados_combinados.append({
                    'data_entrega': status.data_entrega,
                    'instrumento': status.instrumento,
                    'motivo': f'{status.motivo} - {status.observacoes}',
                    'tipo': 'Status'
                })

            # Ordena os dados combinados pela data de entrega
            dados_combinados.sort(key=lambda x: (x['data_entrega'], x['tipo'] == 'Assinatura'))

            wb = load_workbook("Termo de Responsabilidade Equipamentos de Medição.xlsx")
            ws = wb.active

            # Lista para armazenar caminhos de arquivos temporários
            temp_files = []

            if dados_combinados:
                for i, dado in enumerate(dados_combinados):
                    linha_destino = 18 + i

                    # Copia o valor e o estilo da linha 18 para a linha de destino
                    for coluna in range(1, 11):  # A coluna 1 é a A, a coluna 2 é a B, etc.
                        ws.cell(row=linha_destino, column=coluna).font = copy.copy(ws.cell(row=18, column=coluna).font)
                        ws.cell(row=linha_destino, column=coluna).fill = copy.copy(ws.cell(row=18, column=coluna).fill)
                        ws.cell(row=linha_destino, column=coluna).border = copy.copy(ws.cell(row=18, column=coluna).border)
                        ws.cell(row=linha_destino, column=coluna).alignment = copy.copy(ws.cell(row=18, column=coluna).alignment)

                    # Define a altura da linha
                    line_height = ws.row_dimensions[18].height
                    ws.row_dimensions[linha_destino].height = line_height

                    pontos_calibracao = dado['instrumento'].pontos_calibracao.filter()
                    ponto_calibracao_str = ", ".join([ponto.faixa_nominal for ponto in pontos_calibracao])
                    ws.cell(row=linha_destino, column=1, value=dado['data_entrega'].strftime("%d/%m/%Y"))
                    ws.cell(row=linha_destino, column=2, value=dado['instrumento'].tipo_instrumento.nome)
                    ws.cell(row=linha_destino, column=5, value=dado['instrumento'].tag)
                    ws.cell(row=linha_destino, column=7, value=ponto_calibracao_str)
                    ws.cell(row=linha_destino, column=8, value=dado['motivo'])

                    if dado['tipo'] == 'Status':
                        ws.cell(row=linha_destino, column=9, value=dado['data_entrega'].strftime("%d/%m/%Y"))

                    ws.merge_cells(start_row=linha_destino, start_column=2, end_row=linha_destino, end_column=4)
                    ws.merge_cells(start_row=linha_destino, start_column=5, end_row=linha_destino, end_column=6)
                    ws.merge_cells(start_row=linha_destino, start_column=10, end_row=linha_destino, end_column=11)

                    if dado['tipo'] == 'Assinatura' and dado['foto_assinatura']:
                        # Baixa a imagem temporariamente
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
                            file_content = dado['foto_assinatura'].read()
                            temp_file.write(file_content)
                            temp_file_path = temp_file.name
                            temp_files.append(temp_file_path)  # Armazena o caminho do arquivo temporário

                        # Adiciona a imagem ao Excel
                        img = Image(temp_file_path)
                        img.width, img.height = 170, 80  # Ajuste o tamanho conforme necessário
                        img.anchor = f'J{linha_destino}'  # Define a âncora para a célula J
                        ws.add_image(img)

            # Adiciona a última assinatura ao Excel
            if assinatura_funcionario.exists() and assinatura_funcionario.last().foto_assinatura:
                with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as temp_file:
                    file_content = assinatura_funcionario.last().foto_assinatura.read()
                    temp_file.write(file_content)
                    temp_file_path = temp_file.name
                    temp_files.append(temp_file_path)  # Armazena o caminho do arquivo temporário

                # Adiciona a imagem ao Excel
                ultima_foto = Image(temp_file_path)
                ultima_foto.width, ultima_foto.height = 100, 50  # Ajuste o tamanho conforme necessário
                ultima_foto.anchor = 'D12'  # Define a âncora para a célula D12
                ws.add_image(ultima_foto)

            # Preenche os dados do funcionário
            ws['B6'] = funcionario.nome
            ws['H6'] = funcionario.matricula
            ws['H7'] = funcionario.setor.nome
            ws.merge_cells(start_row=12, start_column=3, end_row=12, end_column=5)
            
            # Cria a pasta 'ficha' se não existir
            output_dir = "media/ficha"
            os.makedirs(output_dir, exist_ok=True)

            # Salva o workbook modificado
            output_file_path = os.path.join(output_dir, f"Termo de Responsabilidade Equipamentos de Medição_Atualizado-{funcionario.nome}.xlsx")
            wb.save(output_file_path)
            wb.close()

            # Prepara o arquivo para download
            with open(output_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = f'attachment; filename="Termo de Responsabilidade Equipamentos de Medição_Atualizado-{funcionario.nome}.xlsx"'

            # Remove os arquivos temporários após o salvamento do Excel
            for temp_file_path in temp_files:
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)

            return response
        else:
            return JsonResponse({"message": "Funcionário não encontrado"}, status=404)